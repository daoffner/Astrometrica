import openpyxl
import re
import os
import threading
import time
import glob

def find_orb():
    os.system("./fo mpc.txt")

wb = openpyxl.load_workbook('Possibles 3.xlsx', data_only=True)
sheet = wb.get_sheet_by_name('Sheet1')
x= openpyxl.utils.get_column_letter(sheet.max_column)

#print (ord(x))       Prints the ascii value of the max column; needed for testing
x=str(chr(ord(x)))
y=str(sheet.max_row)
#print(y)               Prints the row size of the sheet; needed for testing
tuple (sheet['A1' : x+y])

path= 'MPC'
dirs = os.listdir("MPC")
mpcArray= []
imageArray= []
i=0

for filename in glob.glob(os.path.join(path, '*.txt')):
    r= open(filename, "r")
    for line in r.readlines():
        mpcArray.append(line.strip())
        i=i+1
    r.close()

f= open("mpc.txt", "w")
g= open("image.txt", "w")
k= 0
size=0

for j in range(int(y)):
        name= sheet['A'+str(j+1)].value
        image= sheet['F'+str(j+1)].value
        #print(re.search(name, mpcArray[i], flags=0))
       regex= re.search(name, mpcArray[i], flags=0)        #Searches for Objects that are in both files
        if regex:
            f.write("     "+mpcArray[i]+'\n')               #Writes a new MPC Report need to create an orbit
            size=size+1
            if(len(imageArray) != 0):
                if(imageArray[k] != image):
                    k=k+1
                    imageArray.append(image)
                    g.write(image+'\n')
            else:
                    imageArray.append(image)
                    g.write(image+'\n')
            mpcArray[i]= "0"

f.close()
g.close()
t = threading.Thread(target= find_orb)
t.daemon= True
t.start()

time.sleep(size/4)
os.system("killall -9 fo")
os.remove('MPCORB.DAT')
os.rename('mpc_fmt.txt','MPCORB.DAT')
os.system('cp -u MPCORB.DAT /mnt/c/Astrometrica/Catalogs')
